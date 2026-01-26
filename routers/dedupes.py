from fastapi import APIRouter,HTTPException,Depends,UploadFile,File,Query,Path
from fastapi import status as http_status
from openpyxl import load_workbook
from typing import Annotated
from datetime import date
import openpyxl
import random
import shutil

import uuid
from io import BytesIO
import re 
import os
import pandas as pd
from pathlib import Path as UploadPath
from typing import List

import string
import time
from typing import Optional
from datetime import datetime
from sqlmodel import select
from sqlmodel.ext.asyncio.session import AsyncSession
from sqlalchemy import insert,func,update,delete

from sqlalchemy.dialects.postgresql import insert as pg_insert

from models.campaigns import Deduped_Campaigns,dedupe_campaigns_tbl
from models.dedupe_keys_table import manual_dedupe_key_tbl

from models.campaign_dedupe import Campaign_Dedupe
from database.master_db_connect import get_async_session
from crud.dedupe_campaigns import (get_dedupe_campaigns_aggregated_count_db,search_cell_number_history_db,search_id_number_history_db,search_dedupe_campaign_by_campaign_name_db)
from crud.campaign_rules import (get_campaign_rule_by_rule_name_db)
from crud.campaigns import get_active_campaign_to_load

from utils.load_campaign_helpers import load_leads_for_campaign

from models.information_table import info_tbl

from models.campaigns import manual_dedupe_keys

from models.dedupe_history_tracker import Dedupe_History_Tracker
from schemas.dedupe_campaigns import SubmitDedupeReturnSchema,DeleteCamapignSchema,UpdateDedupeCampaign
from schemas.dedupes import AddDedupeListResponse,SubmitDedupeReturnResponse,AddManualDedupeResponse,InsertDataDedupeTracker,PaginatedResultsResponse,PaginatedAggregatedDedupeResult

from utils.auth import get_current_user,get_current_active_user
from utils.logger import define_logger
from utils.add_dedupe_list_helpers import add_dedupe_list_helper
from utils.dedupes.submit_return_helpers import update_campaign_dedupe_status,fetch_delete_update_pending_campaign_ids,calculate_ids_campaign_dedupe_with_status_r
from utils.dedupes.manual_dedupe_helpers import insert_campaign_dedupe_batch, insert_manual_dedupe_info_tbl,read_file_into_dict_list,insert_dedupe_data_in_batches
from database.master_db_connect import get_async_session
from database.master_database_prod import get_async_master_prod_session
from utils.load_campaign_helpers import load_leads_for_campaign
from schemas.dedupes import ManualDedupeUploadResponse,SubmitDedupeReturnResponse
from schemas.status_data_routes import InsertStatusDataResponse,InsertEnrichedDataResponse
from crud.dedupe_campaigns import create_manual_dedupe_key

dedupe_logger=define_logger("als_dedupe_campaign_logs","logs/dedupe_route.log")

dedupe_routes=APIRouter(tags=["Dedupes"],prefix="/dedupes")


ALLOWED_EXTS=(".xlsx",".xlsm")
UPLOAD_DIR=UploadPath("uploads/manual_dedupe")
UPLOAD_DIR.mkdir(parents=True,exist_ok=True)

def chunked_manual_dedupe(data: list[dict], size: int = 5000):
    for i in range(0, len(data), size):
        yield data[i : i + size]


def safe_name(name:str)->str:
    return UploadPath(name).name

def make_key(source_name:str)->str:
    suffix = "".join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"{source_name}{suffix}"

@dedupe_routes.post("/add_manual_list2",status_code=http_status.HTTP_200_OK,response_model=ManualDedupeUploadResponse)

async def add_manual_dedupe_list(filename:Optional[str]=None,campaign_name:Optional[str]=None,file:Optional[UploadFile]=File(None),session:AsyncSession=Depends(get_async_session)):
    
    if not campaign_name:
        raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"Missing required query parameter:{campaign_name}")
    if file is None and not filename:

        raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"Provide either query param 'filename' (server file) OR upload a 'file'.")
    temp_saved_path:Optional[Path]=None
    excel_path:str
    key_source_name:str

    if file is not None:
        # Validate uploaded file extension (basic guard)
        original_name = file.filename or "upload.xlsx"

        if not original_name.lower().endswith(ALLOWED_EXTS):
            raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail="Only Excel files (.xlsx, .xlsm) are supported.")
        
        # If user provided filename in query, use it as the server save name,
        # otherwise save using the uploaded original file name.
        save_name = safe_name(filename) if filename else safe_name(original_name)
        # Avoid collisions (append uuid)
        temp_saved_path = UPLOAD_DIR / f"{uuid.uuid4().hex}_{save_name}"
        #this might fail

        try:
            with temp_saved_path.open("wb") as out:
                shutil.copyfileobj(file.file,out)
        finally:
            file.file.close()
        
        excel_path = str(temp_saved_path)

        key_source_name = save_name
    
    else:
        # Using an existing server-side file (filename is a query param)
        assert filename is not None
        if not os.path.isfile(filename):
            raise HTTPException(status_code=http_status.HTTP_404_NOT_FOUND,detail=f"Excel file '{filename}' does not exist on the server.")
        
        if not filename.lower().endswith(ALLOWED_EXTS):
            raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail="Only Excel files (.xlsx, .xlsm) are supported.")
    
        excel_path = filename

        key_source_name = safe_name(filename)
    

    try:
        wb = load_workbook(filename=excel_path, read_only=True, data_only=True)

        rows: list[tuple[str, str]] = [
            (str(id_val), str(cell_val))
            for sheet in wb.worksheets
            for (id_val, cell_val, *_) in sheet.iter_rows(
                min_row=1, min_col=1, max_col=2, values_only=True
            )
            if id_val is not None and cell_val is not None
        ]
    

    except Exception:

        raise HTTPException(status_code=http_status.HTTP_422_UNPROCESSABLE_ENTITY,detail="Could not read the Excel file. It may be corrupted or not a valid workbook.")
    
    finally:
        # cleanup uploaded temp file (we saved it only to parse)
        if temp_saved_path is not None:
            try:
                temp_saved_path.unlink(missing_ok=True)
            except Exception:
                pass
    if not rows:
        return ManualDedupeUploadResponse(success=False,key=None,rows_inserted=0,message="No valid rows found in columns A (id) and B (cell).")
    
     #generate batch key
    
    key=make_key(key_source_name)
    today=date.today()
    #store key in manual_dedupe_key_tbl
    session.add(manual_dedupe_key_tbl(campaign_name=campaign_name,dedupe_key=key))
    #bulk payloads
    dedupe_rows=[
            {
                "cell":r_cell,
                "id":r_id,
                "campaign_name":campaign_name,
                "status":"P",
                "last_used":today,
                "code":key
            }
            for r_id,r_cell in rows
        ]
    
    #prepare bulk payloads on the info_tbl
    info_rows=[
            {
                "cell":r_cell,
                "extra_info":"DEDUPE",
                "norm_cell":r_cell
            }
            for _,r_cell in rows
    ]

    #database operations
    try:
        #insert campaign_dedupe in chunks
        for batch in chunked_manual_dedupe(dedupe_rows,size=5000):
            await session.execute(insert(Campaign_Dedupe).values(batch))
        
        #upsert info_tbl by unique cell
        for batch in chunked_manual_dedupe(info_rows,size=5000):
            stmt=pg_insert(info_tbl).values(batch).on_conflict_do_update(
                index_elements="cell",
                set_={
                    "extra_info": pg_insert(info_tbl).excluded.extra_info,
                    "norm_cell": pg_insert(info_tbl).excluded.norm_cell,
                },
            )

            await session.execute(stmt)
        await session.commit()
        rows_length=len(rows)
        dedupe_logger.info(f"manual dedupe completed with {rows_length} affected")

    except HTTPException:
        raise

    except Exception as e:

        dedupe_logger.exception(f"an exception occurred while adding a manual dedupe:{str(e)}")
        await session.rollback()
        raise HTTPException(status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"Database operations failed")

    return ManualDedupeUploadResponse(
        success=True,
        key=key,
        rows_inserted=rows_length,
        message="Manual dedupe uploaded successfully.",
    )


@dedupe_routes.post("/submit-dedupe-to-als",status_code=http_status.HTTP_200_OK,response_model=SubmitDedupeReturnResponse)

async def submit_dedupe_returned_with_code_and_campaign_name(camp_code:str,dedupe_key:str,file:UploadFile=File(...),session:AsyncSession=Depends(get_async_session),user=Depends(get_current_active_user)):
    
    #guard against empty file name
    if not file.filename:
        raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"No file was uploaded")
    
    #read file
    contents=await file.read()
    #Guard against empty file submission

    if not contents or len(contents)==0:
        raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"Uploaded file is empty")
    
    #split lines
    lines=contents.splitlines()
    total_lines_in_file = len(lines)
    #Guard :file contains only blank lines
    non_empty_lines=[line for line in lines if line.strip()]

    if not non_empty_lines:
        raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"Uploaded file contains no data")
    
    #extract 13 digits ids
    dedupe_ids:List[str]=[]

    for raw in non_empty_lines:
        s=raw.decode("utf-8", errors="ignore").strip()
        if re.fullmatch(r"\d{13}", s):
            dedupe_ids.append(s)
    if not dedupe_ids:
        raise HTTPException(status_code=http_status.HTTP_422_UNPROCESSABLE_ENTITY,detail=f"File does not contain any valid 13-digit ID numbers")
    #remove duplicates while preserving order
    dedupe_ids=list(dict.fromkeys(dedupe_ids))
    #ensure the code exists
    res=await session.execute(select(func.count()).select_from(Campaign_Dedupe).where(Campaign_Dedupe.code==dedupe_key))
    code_count=int(res.scalar() or 0)
    if code_count==0:
        raise HTTPException(status_code=http_status.HTTP_404_NOT_FOUND,detail=f"Dedupe code:{dedupe_key} was not found on the campaign_dedupe table")
    rule = await get_campaign_rule_by_rule_name_db(camp_code, session, user)
    if rule==None:
        raise HTTPException(status_code=http_status.HTTP_404_NOT_FOUND,detail=f"The provided campaign code:{camp_code} does not exist or it's not a dedupe campaign")

    try:
        await session.execute(update(Campaign_Dedupe).where(Campaign_Dedupe.code==dedupe_key).where(Campaign_Dedupe.id.in_(dedupe_ids)).values(status="R"))
        res = await session.execute(select(Campaign_Dedupe.id).where(Campaign_Dedupe.status == "P").where(Campaign_Dedupe.code == dedupe_key))
        removed_ids = [r[0] for r in res.all() if r[0] is not None]
        #delete and update from campaign_dedupe + clear_info
        if removed_ids:
            await session.execute(delete(Campaign_Dedupe).where(Campaign_Dedupe.id.in_(removed_ids)))
            await session.execute(update(info_tbl).where(info_tbl.id.in_(removed_ids)).values(extra_info=None))
        await session.execute(delete(Campaign_Dedupe).where(Campaign_Dedupe.code == "U"))
        await session.commit()
    except HTTPException:
        raise 
    except Exception as e:
        await session.rollback()
        dedupe_logger.exception(f"An exception occurred while processing dedupe operation:{str(e)}")
        raise HTTPException(status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"An internal server error occured, Database operation failed")
    length_of_dedupe_ids=len(dedupe_ids)
    length_of_removed_ids=len(removed_ids)

    return SubmitDedupeReturnResponse(
        success=True,
        campaign_name=camp_code,
        code=dedupe_key,
        total_lines_in_file=total_lines_in_file,
        valid_ids_processed=length_of_dedupe_ids,
        removed_ids_count=length_of_removed_ids,
        message="Dedupe return processed successfully.",
        processed_at=datetime.utcnow(),
    )


@dedupe_routes.post("/add-dedupes-manually",status_code=http_status.HTTP_200_OK)
async def add_dedupes_manually(campaign_name:str=Query(description="Please provide the campaign name"),file:UploadFile=File(...,description="Please provide a file for a manual dedupe"),session:AsyncSession=Depends(get_async_master_prod_session),user=Depends(get_current_user)):
    
    #we are going to use these rows
    all_rows=[]

    file_path=f"temp_{file.filename}"

    #check the extension of the file format being uploaded
    if not file.filename.endswith((".csv",".xlsx",".xls")):
        raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail="Invalid file format")
    try:
        #read the file's content into a BytesIO object(memory file)
        file_content=await file.read()
        file_stream=BytesIO(file_content)

        if file.filename.endswith(".csv"):
            #csv processing logic
            pass

        else:
            #check the campaign codes chiefs??
            workbook=openpyxl.load_workbook(file_stream,read_only=True)

            for sheet in workbook.worksheets:

                extracted_data=[
                  [
                      str(row[0] if row[0] is not None else ""),
                      str(row[1] if row[1] is not None else "")
                  ]
                  for row in sheet.iter_rows(min_col=1,max_col=2,values_only=True)
                  if row[0] is not None or row[1] is not None
                ]
                
                all_rows.extend(extracted_data)
                
            #insert the data into the campaign_dedupe using bulk insert, we still need the campaign name and is_verified field filled
            data_to_insert=[
                {
                    "id":row[0],
                    "cell":row[1],
                    "campaign_codes":campaign_name,
                    "is_verified":True
                }
                for row in all_rows
            ]

            batch_size=1000
            total_rows=len(data_to_insert)

            for i in range(0,total_rows,batch_size):
                batch=data_to_insert[i:i+batch_size]
                insert_stmt=insert(Campaign_Dedupe).values(batch)
                await session.execute(insert_stmt)
            await session.commit()

            dedupe_logger.info(f"user:{user.id} with email:{user.email} uploaded a deduped file for campaign:{campaign_name}")
            return {"message":f"file with name:{file.filename} for campaign:{campaign_name} uploaded successfully"}

    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        dedupe_logger.exception(f"An exception occurred while adding manual dedupe:{e}")
        raise HTTPException(status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"{str(e)}") 


#add dedupe list for a campaign, campaign code is an arguement and return a filename with leads

@dedupe_routes.post("/add-dedupe-list",status_code=http_status.HTTP_200_OK,response_model=AddDedupeListResponse)
async def add_dedupe_list(camp_code:str,session:AsyncSession=Depends(get_async_master_prod_session),user=Depends(get_current_active_user)):
   
    try:
        #fetch a dedupe campaign that matches the given campaign code
        # campaign=session.exec(select(dedupe_campaigns_tbl).where(Deduped_Campaigns.camp_code==camp_code)).first()
        #this function must return the campaign code for the dedupe campaign, use that code to find the matching rule
        campaign=await get_active_campaign_to_load(camp_code,session,user)

        #raise an exception if the dedupe campaign does not exist prompting the user to create a rule for that campaign
        if campaign==None:
            dedupe_logger.info(f"dedupe campaign with code:{camp_code} does not exist or it is not a dedupe campaign,requested by user:{user.id} with email:{user.email}")
            raise HTTPException(status_code=http_status.HTTP_404_NOT_FOUND,detail=f"Campaign with campaign code:{camp_code} does not exist or it is not a dedupe campaign")
        #search the dedupe campaigns rules tbl using the campaign code
    

        #find the dedupe campaign rule
        campaign_rule=await get_campaign_rule_by_rule_name_db(camp_code,session,user)
        if campaign_rule==None:
            dedupe_logger.info(f'campaign rule with rule name:{camp_code} does not exist')
            raise HTTPException(status_code=http_status.HTTP_404_NOT_FOUND,detail=f"campaign rule with rule name:{campaign_rule} does not exist")
        #now we have the actual list here,
        results=await load_leads_for_campaign(campaign_rule.rule_name,session)

        #get leads length
        leads_length=len(results)
        if leads_length==0:
            dedupe_logger.info(f"No leads found for dedupe campaign:{camp_code}")
            raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"No leads found for the dedupe campaign:{camp_code}")
        #set the today's list
        todaysdate=datetime.today().strftime('%Y-%m-%d')

        filename=camp_code + '-'+ todaysdate
        #insert leads inside the campaign_dedupe table
        suffix=''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(4))
        #create file key
        key=filename + suffix
        #prepare list of tuple to insert into the table campaign_dedupe table
        keys_to_keep=['id','cell']
        
        insert_list_to_campaign_dedupe_table=[{k: d[k] for k in keys_to_keep} for d in results]
        
        #convert the list of dictionaries into a list of tuples
        dedupe_values={"campaign_code":camp_code,"status":'P',"key":key}
        list_to_insert_to_campaign_dedupe=[{**d,**dedupe_values} for d in insert_list_to_campaign_dedupe_table]
        #convert the entries to a list of tuples to insert to the database
        list_to_insert_to_db=[tuple(d.values()) for d in list_to_insert_to_campaign_dedupe]
        #insert the add dedupe list into the db
        insert_result=await add_dedupe_list_helper(session,list_to_insert_to_db,batch_size=1000)
        
        if '/' in filename:
            new_string_name=filename.replace("/","-")
        else:
            new_string_name=filename

        #convert the list to a dataframe, using list of tuples
        converted_list=[tuple(d.values()) for d in results]
        #convert the list of dictionaries to list of tuples
        df=pd.DataFrame(converted_list)
        #convert the dataframe to an excel file
        df.to_excel(new_string_name + '.xlsx',index=False)
        await session.commit()
        dedupe_logger.info(f"user:{user.id} with email:{user.email} added a dedupe list")
        return AddDedupeListResponse(FileName=new_string_name + '.txt',TotalRecordsInserted=insert_result["total_inserted"],TotalBatches=insert_result["total_batches"],TotalBatchedTime=insert_result["batch_times"],TotalTime=insert_result["total_time"])
        
    
    except Exception as e:
        dedupe_logger.exception(f"Internal server error occurred while adding a dedupe list for campaign {camp_code}:{e}")
        await session.rollback()
        raise HTTPException(status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"Internal server error occured while adding a dedupe list for campaign:{camp_code}")

#submit dedupe return route
@dedupe_routes.post("/submit-dedupe-return",status_code=http_status.HTTP_200_OK,description="Submit file returned from dedupe,this route make sure that only approved cell/id numbers are used for the dedupe campaign and all unapproved numbers are released from being flagged",response_model=SubmitDedupeReturnResponse)
async def submit_dedupe_return(data:SubmitDedupeReturnSchema,dedupe_file:UploadFile=File(...,description="File prepared for manual dedupe"),session:AsyncSession=Depends(get_async_session),user=Depends(get_current_active_user)):
    try:
        id_pattern=re.compile(r"^\d{13}$")
        file_contents=await dedupe_file.read()
        if not file_contents:
            dedupe_logger.info(f"user {user.id} with email:{user.email} uploaded an empty file")
            raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"Empty file uploaded")
        text=file_contents.decode("utf-8",errors="strict")
        
        result_count=await calculate_ids_campaign_dedupe_with_status_r(session,data.dedupe_code)

        if result_count==0:
            dedupe_logger.info(f"No results associated with the dedupe key:{data.dedupe_code}")
            raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"No results associated with the dedupe key:{data.code}")
        
        dedupe_ids=[line for line in text.splitlines() if id_pattern.match(line)] 
        updated_count=await update_campaign_dedupe_status(dedupe_ids,data.dedupe_code,session,user)
        results=await fetch_delete_update_pending_campaign_ids(data.dedupe_code,session,user)
        await session.commit()

        return SubmitDedupeReturnResponse(success=True,
                                          number_of_records_from_campaign_dedupe_table_with_mathing_dedupe_code=result_count,
                                          update_campaign_dedupe_table_with_return_status_with_id_numbers=updated_count,
                                          number_of_records_with_status_process_from_campaign_dedupe_table=results["retrieved_pending_ids"],
                                          number_of_deleted_records_from_campaign_dedupe_table=results["deleted_ids_from_campaign_dedupe_table"],
                                          number_of_records_updated_on_the_info_table=results["updated_ids_from_info_tbl"],
                                          number_of_records_deleted_with_status_updated_on_the_campaign_dedupe_table=results["deleted_stmt_from_campaign_dedupe"]
                                         )
    


    
    except Exception as e:
        dedupe_logger.exception(f"an internal server error occurred while submitting dedupe by:{user.id} with email:{user.email} {e}")
        raise HTTPException(status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"An internal server error occurred while submitting dedupe with dedupe key:{data.code}")


@dedupe_routes.post("/add-manual-dedupe-list2",status_code=http_status.HTTP_201_CREATED,response_model=AddManualDedupeResponse)
async def add_manual_dedupe_list2(filename:Annotated[UploadFile,File(description="Upload excel file with cell numbers")],camp_code:str=Query(description="Campaign Code"),user=Depends(get_current_active_user),session:AsyncSession=Depends(get_async_master_prod_session)): 
    
    try:
        #read the file safely
        file_content=await filename.read()
        if not file_content:
            dedupe_logger.info(f"an empty file was uploaded")
            raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"Empty file uploaded")
        #validate the id number and cell number
        za_id_pattern=re.compile(r"^\d{13}$")

        za_cell_pattern=re.compile(r"^0[678]\d{8}$")
        #check the file format
        #this could have been in a method for clean reading

        try:
            wb=openpyxl.load_workbook(BytesIO(file_content),read_only=True)
        except Exception as e:
            dedupe_logger.exception(f"Error occurred while reading from the workbook",e)
            raise HTTPException(status_code=http_status.HTTP_422_UNPROCESSABLE_ENTITY,detail=f"the file uploaded is not a valid or readable Excel workbook (.xlsx)") 
        rows=[]

        try:

            for sheet in wb.worksheets:
                #skip header row
                for id_num,cell_num in sheet.iter_rows(min_row=2,max_col=2,values_only=True):
                    if id_num is None or cell_num is None:
                        continue
                    id_num_str=str(id_num).strip()
                    cell_str=str(cell_num).strip()
                    if not za_id_pattern.match(id_num_str):
                        continue
                    if not za_cell_pattern.match(cell_str):
                        continue
                    #line=[id_num_str,cell_num]
                    rows.append((id_num_str,cell_str))
        finally:
            wb.close()
        if not rows:
            dedupe_logger.info(f"An error occurred while loading data into the rows list")
            raise HTTPException(status_code=http_status.HTTP_400_BAD_REQUEST,detail=f"No data could be extracted")
        suffix=''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(4))
        key=filename.filename + suffix
        #data to insert in the campaign_dedupe table
        data=[(r[0],r[1],camp_code,'P',key) for r in rows]
        #call the insert method
        inserted_rows_campaign_dedupe_table=await insert_campaign_dedupe_batch(session,data,user)
        #insert into the info_tbl
        info_tbl_data=[(r[1],'DEDUPE') for r in rows]
        inserted_rows_on_info_tbl=await insert_manual_dedupe_info_tbl(session,info_tbl_data,user)
        #store the dedupe key somewhere
        dedupe_key=await create_manual_dedupe_key(session,rule_name=camp_code,dedupe_key=key,number_of_leads=len(data),user=user)
        #commit everything in one transaction fool
        await session.commit()
        dedupe_logger.info(f"user:{user.id} with email:{user.email} inserted {inserted_rows_campaign_dedupe_table} records into the campaign_dedupe table and inserted:{inserted_rows_on_info_tbl}")
        return AddManualDedupeResponse(success=True,campaign_dedupe_records=inserted_rows_campaign_dedupe_table,info_table_records=inserted_rows_on_info_tbl,key=dedupe_key.dedupe_key)
    except HTTPException:
        raise 

    except Exception as e:
        dedupe_logger.exception(f"An internal server error while adding manual dedupe:{e}")
        raise HTTPException(status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while adding a manual dedupe")
    

@dedupe_routes.post("/load_dedupe_tracker",status_code=http_status.HTTP_200_OK,description="upload a csv or excel file with cell numbers and/or id numbers along with the dedupe campaign name and code",response_model=InsertDataDedupeTracker)
async def upload_dedupe_campaign_records(campaign_name:str=Query(...,description="campaign name with records"),camp_code:str=Query(...,description="camp_code for the campaign"),session:AsyncSession=Depends(get_async_master_prod_session),file:UploadFile=File(...,description="Dedupe file with id numbers or cell phones that must be a csv file or excel file"),user=Depends(get_current_active_user)):
    
    try:
        #extract the information on the file
        data=await read_file_into_dict_list(file,campaign_name=campaign_name,camp_code=camp_code)
        #load the table tracker with the extracted data
        result=await insert_dedupe_data_in_batches(session,data)
        return InsertDataDedupeTracker(message=result["message"],number_of_batches=result["total_batches_processed"],number_of_records=result["total_records_inserted"])
    except Exception as e:
        dedupe_logger.exception(f"An exception occurred while uploading a dedupe file:{e}")
        raise HTTPException(status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"An internal server error occurred while uploading a dedupe file for campaign:{campaign_name} with campaign code:{camp_code}")


#get all records for each campaign
@dedupe_routes.get("/search/campaigns",status_code=http_status.HTTP_200_OK,description="get all records for each campaign",response_model=PaginatedResultsResponse)
async def search_dedupe_campaign_by_campaign_name(campaign_name:str=Query(...,description="Provide the campaign name"),page:int=Query(1,ge=1,description="This value should be greater than one"),page_size:int=Query(10,ge=1),session:AsyncSession=Depends(get_async_master_prod_session),user=Depends(get_current_active_user)):
    return await search_dedupe_campaign_by_campaign_name_db(campaign_name,page,page_size,session)

#search by id
@dedupe_routes.get("/search/id",status_code=http_status.HTTP_200_OK,description="Search the id history",response_model=PaginatedResultsResponse)
async def search_by_id_number(id_number:str=Query(None,description="Enter the id number to search it's history"),page:int=Query(1,ge=1),page_size:int=Query(10,ge=1),session:AsyncSession=Depends(get_async_master_prod_session),user=Depends(get_current_active_user)):
    return await search_id_number_history_db(id_number,page,page_size,session,user)

#search by cell number
@dedupe_routes.get("/search/cell",status_code=http_status.HTTP_200_OK,description="Search the history of the cell number",response_model=PaginatedResultsResponse)
async def search_cell_phone_number_history(cell:str=Query(...,description="Enter cell number"),campaign_name:Optional[str]=Query(None,description="Enter the campaign name if available"),page:int=Query(1,ge=1),page_size:int=Query(10,ge=1),session:AsyncSession=Depends(get_async_master_prod_session),user=Depends(get_current_active_user)):
    return await search_cell_number_history_db(cell,page,page_size,session,user,campaign_name)


#show aggregated count of entries for each campaign by campaign name

@dedupe_routes.get("/campaigns/campaigns_aggregated_count",status_code=http_status.HTTP_200_OK,response_model=PaginatedAggregatedDedupeResult)
async def get_campaign_aggregated_count(page:int=Query(1,ge=1),page_size:int=Query(10,ge=1,le=100),session:AsyncSession=Depends(get_async_master_prod_session),user=Depends(get_current_active_user)):
    return await get_dedupe_campaigns_aggregated_count_db(page,page_size,session,user)


@dedupe_routes.get("/campaigns/search-campaign",status_code=http_status.HTTP_200_OK)
async def search_dedupe_campaign(campaign_code:str=Query(...,description="Provide the campaign code to retrieve from the database"),page:int=Query(1,ge=1),page_size:int=Query(10,ge=1,le=100),session:AsyncSession=Depends(get_async_master_prod_session),user=Depends(get_current_active_user)):
    return True
