from openpyxl import load_workbook
from sqlalchemy import text

from sqlalchemy.ext.asyncio.session import AsyncSession

async def load_excel_into_info_tbl(session: AsyncSession, path: str, batch_size: int = 5000) -> int:
    
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active  # or wb["SheetName"]

    inserted = 0
    batch: list[dict] = []

    # Example: first row is headers
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)

    # Adjust indexes based on your Excel columns
    # Example assumes: col A = cell, col B = extra_info
    for row in rows_iter:
        cell_value = row[0]
        extra_info = row[1]

        # skip empty rows
        if cell_value is None:
            continue

        batch.append({"cell": str(cell_value).strip(), "extra_info": (str(extra_info).strip() if extra_info else None)})

        if len(batch) >= batch_size:
            await session.execute(text(INSERT_INFO_TBL), batch)
            await session.commit()  # commit per batch = safer + less memory
            inserted += len(batch)
            batch.clear()

    # last batch
    if batch:
        await session.execute(text(INSERT_INFO_TBL), batch)
        await session.commit()
        inserted += len(batch)

    wb.close()
    return inserted